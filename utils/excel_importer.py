import argparse
import asyncio
from pathlib import Path
from typing import Any
import openpyxl
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from config.db import AsyncSessionLocal
from utils.security import hash_password
from utils.uuid_utils import generate_uuid7


def clean_val(val: Any) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    if s.endswith('.0') and s[:-2].isdigit():
        return s[:-2]
    return s if s else None


def clean_int(val: Any) -> int | None:
    if val is None:
        return None
    s = str(val).strip()
    if s.endswith('.0'):
        s = s[:-2]
    try:
        return int(s)
    except ValueError:
        return None


class ExcelMemberImporter:
    """Wrapper for parsing and importing KSM AIoT members from Excel to PostgreSQL."""

    @staticmethod
    def parse_excel(file_path: str | Path | Any, sheet_name: str = "Database Anggota") -> list[dict]:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' tidak ditemukan di file Excel. Sheets yang tersedia: {wb.sheetnames}")

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        raw_headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]

        members = []
        for idx, row in enumerate(rows[1:], start=1):
            if not any(row):
                continue

            row_dict = dict(zip(raw_headers, row))

            student_id = clean_val(row_dict.get("NIM"))
            full_name = clean_val(row_dict.get("Nama Lengkap")) or clean_val(row_dict.get("Nama Lengkap "))
            if not student_id or not full_name:
                continue

            # Standardize division
            raw_division = clean_val(row_dict.get("Divisi")) or "BPH"
            division = raw_division
            if "Akademik" in raw_division or "Riset" in raw_division:
                division = "Akademik & Riset"
            elif "Sumber Daya" in raw_division or "PSDM" in raw_division:
                division = "Pengembangan SDM"
            elif "Masyarakat" in raw_division or "Multimedia" in raw_division or "Humas" in raw_division:
                division = "Humas & Multimedia"
            elif "BPH" in raw_division or not raw_division:
                division = "BPH"

            # Standardize role / jabatan
            role = clean_val(row_dict.get("Jabatan")) or "Anggota"

            # Standardize intaking period
            intake_period = clean_val(row_dict.get("Tahun Masuk KSM")) or "2026"

            member_record = {
                "sequence_index": idx,
                "student_id": student_id,
                "full_name": full_name,
                "program_of_study": clean_val(row_dict.get("Program Studi")) or "-",
                "semester": clean_int(row_dict.get("Semester")),
                "email": clean_val(row_dict.get("Email Pribadi")) or f"{student_id}@mahasiswa.upnvj.ac.id",
                "contact_info": clean_val(row_dict.get("No WhatsApp")) or clean_val(row_dict.get("No WhatsApp ")),
                "domicile_city": clean_val(row_dict.get("Kota Domisili")) or clean_val(row_dict.get("Kota Domisili ")),
                "division": division,
                "role": role,
                "intake_period": intake_period,
                "interest_track": clean_val(row_dict.get("Bidang apa yang ingin kamu eksplorasi lebih jauh di KSM ini? "))
                                 or clean_val(row_dict.get("Keahlian/Fokus utama kamu saat ini? ")),
                "focus_expertise": clean_val(row_dict.get("Keahlian/Fokus utama kamu saat ini? ")),
                "exploration_field": clean_val(row_dict.get("Bidang apa yang ingin kamu eksplorasi lebih jauh di KSM ini? ")),
                "field_reason": clean_val(row_dict.get("Alasan memilih bidang tersebut  ")),
                "programming_languages": clean_val(row_dict.get("Bahasa pemrograman yang dikuasai  ")),
                "tools_frameworks": clean_val(row_dict.get("Tools / Framework yang pernah digunakan  ")),
                "project_experience": clean_val(row_dict.get("Pernah mengerjakan project? Jelaskan project tersebut")),
                "hackathon_experience": clean_val(row_dict.get("Pernah ikut lomba / hackathon?")),
                "portfolio_url": clean_val(row_dict.get("Link Portofolio / Github / Linkedin")),
                "routine_commitment": clean_val(row_dict.get("Bersedia mengikuti kegiatan rutin?  ")),
                "weekly_free_time": clean_val(row_dict.get("Estimasi waktu luang per minggu (jam)")),
                "other_activities": clean_val(row_dict.get("Kesibukan lain")),
                "discord_id": clean_val(row_dict.get("Akun Discord (ID)")),
                "registration_timestamp": clean_val(row_dict.get("Timestamp")),
                "status": clean_val(row_dict.get("Status Anggota")) or "Aktif",
                "join_date": "15/09/2023" if "2023" in intake_period else "20/09/2024" if "2024" in intake_period else "15/01/2026",
            }
            members.append(member_record)

        return members

    @staticmethod
    async def import_to_database(db: AsyncSession, members: list[dict], default_password: str = "aiotupnvj2026") -> dict:
        """Insert or update all members and sync login accounts in PostgreSQL using raw SQL."""
        default_hash = hash_password(default_password)
        imported_count = 0
        user_synced_count = 0

        # Pre-fetch existing members to maintain consistent member_id or assign new ones
        existing_result = await db.execute(text("SELECT id, student_id, member_id FROM members"))
        existing_by_student = {r["student_id"]: r["member_id"] for r in existing_result.mappings().all()}
        used_member_ids = set(existing_by_student.values())

        for idx, m in enumerate(members, start=1):
            student_id = m["student_id"]
            
            # Determine unique member_id
            if student_id in existing_by_student:
                member_id = existing_by_student[student_id]
            else:
                candidate_id = f"AIOT-2026-{str(idx).zfill(3)}"
                counter = idx
                while candidate_id in used_member_ids:
                    counter += 1
                    candidate_id = f"AIOT-2026-{str(counter).zfill(3)}"
                member_id = candidate_id
                used_member_ids.add(member_id)

            # 1. Upsert Member using raw SQL
            m_uuid = generate_uuid7()
            member_stmt = text(
                """
                INSERT INTO members (
                    id, member_id, student_id, full_name, program_of_study, semester, email, contact_info, domicile_city,
                    division, role, intake_period, interest_track, focus_expertise, exploration_field, field_reason,
                    programming_languages, tools_frameworks, project_experience, hackathon_experience, portfolio_url,
                    routine_commitment, weekly_free_time, other_activities, discord_id, registration_timestamp,
                    avatar, status, join_date, created_at
                )
                VALUES (
                    :id, :member_id, :student_id, :full_name, :program_of_study, :semester, :email, :contact_info, :domicile_city,
                    :division, :role, :intake_period, :interest_track, :focus_expertise, :exploration_field, :field_reason,
                    :programming_languages, :tools_frameworks, :project_experience, :hackathon_experience, :portfolio_url,
                    :routine_commitment, :weekly_free_time, :other_activities, :discord_id, :registration_timestamp,
                    :avatar, :status, :join_date, NOW()
                )
                ON CONFLICT (student_id) DO UPDATE SET
                    full_name = EXCLUDED.full_name,
                    program_of_study = EXCLUDED.program_of_study,
                    semester = EXCLUDED.semester,
                    email = EXCLUDED.email,
                    contact_info = EXCLUDED.contact_info,
                    domicile_city = EXCLUDED.domicile_city,
                    division = EXCLUDED.division,
                    role = EXCLUDED.role,
                    intake_period = EXCLUDED.intake_period,
                    interest_track = EXCLUDED.interest_track,
                    focus_expertise = EXCLUDED.focus_expertise,
                    exploration_field = EXCLUDED.exploration_field,
                    field_reason = EXCLUDED.field_reason,
                    programming_languages = EXCLUDED.programming_languages,
                    tools_frameworks = EXCLUDED.tools_frameworks,
                    project_experience = EXCLUDED.project_experience,
                    hackathon_experience = EXCLUDED.hackathon_experience,
                    portfolio_url = EXCLUDED.portfolio_url,
                    routine_commitment = EXCLUDED.routine_commitment,
                    weekly_free_time = EXCLUDED.weekly_free_time,
                    other_activities = EXCLUDED.other_activities,
                    discord_id = EXCLUDED.discord_id,
                    registration_timestamp = EXCLUDED.registration_timestamp,
                    status = EXCLUDED.status,
                    join_date = EXCLUDED.join_date
                """
            )
            await db.execute(member_stmt, {
                "id": m_uuid,
                "member_id": member_id,
                "student_id": student_id,
                "full_name": m["full_name"],
                "program_of_study": m["program_of_study"],
                "semester": m["semester"],
                "email": m["email"],
                "contact_info": m["contact_info"],
                "domicile_city": m["domicile_city"],
                "division": m["division"],
                "role": m["role"],
                "intake_period": m["intake_period"],
                "interest_track": m["interest_track"],
                "focus_expertise": m["focus_expertise"],
                "exploration_field": m["exploration_field"],
                "field_reason": m["field_reason"],
                "programming_languages": m["programming_languages"],
                "tools_frameworks": m["tools_frameworks"],
                "project_experience": m["project_experience"],
                "hackathon_experience": m["hackathon_experience"],
                "portfolio_url": m["portfolio_url"],
                "routine_commitment": m["routine_commitment"],
                "weekly_free_time": m["weekly_free_time"],
                "other_activities": m["other_activities"],
                "discord_id": m["discord_id"],
                "registration_timestamp": m["registration_timestamp"],
                "avatar": None,
                "status": m["status"],
                "join_date": m["join_date"],
            })
            imported_count += 1

            # 2. Upsert corresponding User login account
            user_role = "PENGURUS"
            if "Ketua" in m["role"] and "Wakil" not in m["role"]:
                user_role = "SUPERADMIN"
            elif "Wakil" in m["role"] or "Sekretaris" in m["role"] or "Bendahara" in m["role"]:
                user_role = "ADMIN_BPH"
            elif "Anggota" in m["role"]:
                user_role = "ANGGOTA"

            user_stmt = text(
                """
                INSERT INTO users (id, student_id, full_name, email, hashed_password, role, division, avatar, is_active, created_at)
                VALUES (:id, :student_id, :full_name, :email, :hashed_password, :role, :division, :avatar, true, NOW())
                ON CONFLICT (student_id) DO UPDATE SET
                    full_name = EXCLUDED.full_name,
                    email = EXCLUDED.email,
                    role = EXCLUDED.role,
                    division = EXCLUDED.division
                """
            )
            await db.execute(user_stmt, {
                "id": generate_uuid7(),
                "student_id": student_id,
                "full_name": m["full_name"],
                "email": m["email"],
                "hashed_password": default_hash,
                "role": user_role,
                "division": m["division"],
                "avatar": None,
            })
            user_synced_count += 1

        await db.commit()
        return {
            "status": "success",
            "members_imported": imported_count,
            "users_synced": user_synced_count,
        }


async def main():
    parser = argparse.ArgumentParser(description="Import members from Excel into PostgreSQL")
    parser.add_argument("--file", default="KSM AI IoT.xlsx", help="Path to Excel file")
    parser.add_argument("--sheet", default="Database Anggota", help="Sheet name")
    args = parser.parse_args()

    file_path = Path(args.file)
    if not file_path.exists():
        print(f"Error: File {file_path} not found!")
        return

    print(f"Parsing '{args.sheet}' from {file_path}...")
    members = ExcelMemberImporter.parse_excel(file_path, sheet_name=args.sheet)
    print(f"Parsed {len(members)} member entries.")

    async with AsyncSessionLocal() as session:
        result = await ExcelMemberImporter.import_to_database(session, members)
        print(f"Database sync result: {result}")


if __name__ == "__main__":
    asyncio.run(main())
